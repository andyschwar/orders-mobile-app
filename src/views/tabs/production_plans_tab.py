from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTableWidget, 
    QTableWidgetItem, QLabel, QLineEdit, QComboBox, QSpinBox,
    QDateEdit, QMessageBox, QDialog, QFormLayout, QDialogButtonBox,
    QHeaderView, QCheckBox, QGroupBox, QRadioButton, QButtonGroup,
    QTextEdit, QSplitter, QFileDialog, QApplication
)
from PyQt6.QtCore import Qt, pyqtSignal, QDate
from PyQt6.QtPrintSupport import QPrinter
from PyQt6.QtGui import QFont, QPainter, QTextDocument, QColor, QBrush
from sqlalchemy.orm import Session
from datetime import datetime, date, timedelta
from utils.permissions import get_permissions_manager
import pandas as pd
import os

class ProductionPlanDialog(QDialog):
    def __init__(self, plan_type: str, session: Session, parent=None):
        super().__init__(parent)
        self.plan_type = plan_type
        self.session = session
        self.setup_ui()
        
    def setup_ui(self):
        plan_type_names = {
            'type1': 'Plan Type 1 - By Customer, Order, Delivery Date',
            'type2': 'Plan Type 2 - By Customer, Order, Delivery Date, Order Date, Product',
            'type3': 'Plan Type 3 - Production Summary by Product, Month, Surface Treatment'
        }
        
        self.setWindowTitle(plan_type_names.get(self.plan_type, f"Production Plan {self.plan_type}"))
        self.setModal(True)
        self.resize(600, 500)  # Make dialog wider
        
        layout = QFormLayout()
        
        # Customer selection
        self.customer_combo = QComboBox()
        self.customer_combo.setMinimumWidth(300)  # Make wider
        self.load_customers()
        layout.addRow("Customer:", self.customer_combo)
        
        # Order selection (for type 1 and 2)
        if self.plan_type in ['type1', 'type2']:
            self.order_combo = QComboBox()
            self.order_combo.setMinimumWidth(300)  # Make wider
            self.load_orders()
            layout.addRow("Order:", self.order_combo)
            
            # Connect customer selection to reload orders
            self.customer_combo.currentIndexChanged.connect(self.load_orders)
        
        # Delivery date selection (for type 1)
        if self.plan_type == 'type1':
            delivery_date_label = QLabel("Specific Delivery Date:")
            delivery_date_layout = QHBoxLayout()
            
            self.delivery_date_all = QCheckBox("All dates")
            self.delivery_date_all.setChecked(True)
            delivery_date_layout.addWidget(self.delivery_date_all)
            
            self.delivery_date_combo = QComboBox()
            self.delivery_date_combo.setMinimumWidth(200)
            self.delivery_date_combo.setEnabled(False)  # Disabled when "all" is checked
            delivery_date_layout.addWidget(self.delivery_date_combo)
            
            # Connect checkbox to enable/disable combo
            self.delivery_date_all.toggled.connect(lambda checked: self.delivery_date_combo.setEnabled(not checked))
            
            # Connect customer and order selection to load delivery dates
            self.customer_combo.currentIndexChanged.connect(self.load_delivery_dates)
            self.order_combo.currentIndexChanged.connect(self.load_delivery_dates)
            
            layout.addRow(delivery_date_label, delivery_date_layout)
        
        # Delivery date (month/year) - allow multiple selection
        delivery_label = QLabel("Delivery Date (Month/Year):")
        delivery_layout = QVBoxLayout()
        
        # Month selection - allow multiple
        month_layout = QHBoxLayout()
        month_layout.addWidget(QLabel("Months:"))
        self.delivery_months = []
        for i in range(1, 13):
            checkbox = QCheckBox(f"{i:02d}")
            self.delivery_months.append(checkbox)
            month_layout.addWidget(checkbox)
        delivery_layout.addLayout(month_layout)
        
        # Year selection
        year_layout = QHBoxLayout()
        year_layout.addWidget(QLabel("Year:"))
        self.delivery_year = QComboBox()
        self.delivery_year.setMinimumWidth(150)  # Make wider
        current_year = datetime.now().year
        for year in range(current_year - 1, current_year + 3):
            self.delivery_year.addItem(str(year), year)
        self.delivery_year.setCurrentText(str(current_year))
        year_layout.addWidget(self.delivery_year)
        delivery_layout.addLayout(year_layout)
        
        layout.addRow(delivery_label, delivery_layout)
        
        # Order date (for type 2) - allow "all" option
        if self.plan_type == 'type2':
            order_date_label = QLabel("Order Date:")
            order_date_layout = QHBoxLayout()
            
            self.order_date_all = QCheckBox("All dates")
            self.order_date_all.setChecked(True)
            order_date_layout.addWidget(self.order_date_all)
            
            self.order_date = QDateEdit()
            self.order_date.setDate(QDate.currentDate())
            self.order_date.setCalendarPopup(True)
            self.order_date.setMinimumWidth(200)  # Make wider
            self.order_date.setEnabled(False)  # Disabled when "all" is checked
            order_date_layout.addWidget(self.order_date)
            
            # Connect checkbox to enable/disable date picker
            self.order_date_all.toggled.connect(lambda checked: self.order_date.setEnabled(not checked))
            
            layout.addRow(order_date_label, order_date_layout)
            
            # Product selection
            self.product_combo = QComboBox()
            self.product_combo.setMinimumWidth(300)  # Make wider
            self.load_products()
            layout.addRow("Product:", self.product_combo)
        
        # Surface treatment (for type 3) - allow "all" option
        if self.plan_type == 'type3':
            surface_label = QLabel("Surface Treatment:")
            surface_layout = QHBoxLayout()
            
            self.surface_all = QCheckBox("All treatments")
            self.surface_all.setChecked(True)
            surface_layout.addWidget(self.surface_all)
            
            self.surface_treatment = QComboBox()
            self.surface_treatment.addItems(["KATAFOREZA", "FOSFAT", "ZINEK"])
            self.surface_treatment.setMinimumWidth(200)  # Make wider
            self.surface_treatment.setEnabled(False)  # Disabled when "all" is checked
            surface_layout.addWidget(self.surface_treatment)
            
            # Connect checkbox to enable/disable combo
            self.surface_all.toggled.connect(lambda checked: self.surface_treatment.setEnabled(not checked))
            
            layout.addRow(surface_label, surface_layout)
        
        # Buttons
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addRow("", button_box)
        
        self.setLayout(layout)
    
    def load_customers(self):
        from models.database import Customer
        customers = self.session.query(Customer).order_by(Customer.name_index).all()
        self.customer_combo.clear()
        self.customer_combo.addItem("All Customers", None)
        for customer in customers:
            self.customer_combo.addItem(f"{customer.name_index} - {customer.name}", customer.id)
    
    def load_orders(self):
        from models.database import Order
        customer_id = self.customer_combo.currentData()
        if customer_id:
            orders = self.session.query(Order).filter(Order.customer_id == customer_id).order_by(Order.order_number).all()
        else:
            orders = []
        
        self.order_combo.clear()
        self.order_combo.addItem("All Orders", None)
        for order in orders:
            self.order_combo.addItem(order.order_number, order.id)
    
    def load_products(self):
        from models.database import Product
        products = self.session.query(Product).order_by(Product.name).all()
        self.product_combo.clear()
        self.product_combo.addItem("All Products", None)
        for product in products:
            self.product_combo.addItem(product.name, product.id)
    
    def load_delivery_dates(self):
        """Load available delivery dates for the selected customer/order"""
        from models.database import OrderItem, Order
        
        # Get selected customer and order
        customer_id = self.customer_combo.currentData()
        order_id = self.order_combo.currentData()
        
        # Build query to get unique delivery dates
        query = self.session.query(OrderItem.delivery_date)
        
        if customer_id:
            query = query.join(Order).filter(Order.customer_id == customer_id)
        
        if order_id:
            query = query.filter(OrderItem.order_id == order_id)
        
        # Get unique dates and filter for undelivered items
        query = query.filter(OrderItem.quantity > OrderItem.delivered_quantity)
        query = query.distinct().order_by(OrderItem.delivery_date)
        
        delivery_dates = query.all()
        
        # Populate the combo box
        self.delivery_date_combo.clear()
        for date_obj in delivery_dates:
            if date_obj.delivery_date:
                date_str = date_obj.delivery_date.strftime('%Y-%m-%d')
                self.delivery_date_combo.addItem(date_str, date_obj.delivery_date)
    
    def get_params(self):
        params = {
            'customer_id': self.customer_combo.currentData(),
            'delivery_months': [i+1 for i, month in enumerate(self.delivery_months) if month.isChecked()],
            'delivery_year': self.delivery_year.currentData()
        }
        
        if self.plan_type in ['type1', 'type2']:
            params['order_id'] = self.order_combo.currentData()
        
        if self.plan_type == 'type1':
            if self.delivery_date_all.isChecked():
                params['delivery_date'] = None  # "All dates"
            else:
                params['delivery_date'] = self.delivery_date_combo.currentData()
        
        if self.plan_type == 'type2':
            if self.order_date_all.isChecked():
                params['order_date'] = None  # "All dates"
            else:
                params['order_date'] = self.order_date.date().toPyDate()
            params['product_id'] = self.product_combo.currentData()
        
        if self.plan_type == 'type3':
            if self.surface_all.isChecked():
                params['surface_treatment'] = None  # "All treatments"
            else:
                params['surface_treatment'] = self.surface_treatment.currentText()
        
        return params

class ProductionPlansTab(QWidget):
    plan_updated = pyqtSignal()
    
    def __init__(self, session: Session, user=None, parent=None):
        super().__init__(parent)
        self.session = session
        self.user = user
        self.permissions_manager = get_permissions_manager()
        self.current_plan_data = None
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        
        # Create toolbar
        toolbar = QHBoxLayout()
        
        # Plan type selection
        plan_label = QLabel("Plan Type:")
        self.plan_type_group = QButtonGroup()
        self.plan_type_1 = QRadioButton("Type 1")
        self.plan_type_2 = QRadioButton("Type 2")
        self.plan_type_3 = QRadioButton("Type 3")
        self.plan_type_1.setChecked(True)
        
        self.plan_type_group.addButton(self.plan_type_1)
        self.plan_type_group.addButton(self.plan_type_2)
        self.plan_type_group.addButton(self.plan_type_3)
        
        toolbar.addWidget(plan_label)
        toolbar.addWidget(self.plan_type_1)
        toolbar.addWidget(self.plan_type_2)
        toolbar.addWidget(self.plan_type_3)
        
        # Generate button
        generate_button = QPushButton("Generate Plan")
        generate_button.clicked.connect(self.generate_plan)
        toolbar.addWidget(generate_button)
        
        # Export button
        export_button = QPushButton("Export to Excel")
        export_button.clicked.connect(self.export_plan)
        toolbar.addWidget(export_button)
        
        toolbar.addStretch()
        layout.addLayout(toolbar)
        
        # Plan description
        self.description_label = QLabel("Type 1: List of undelivered items by Customer, Order, Delivery Date")
        self.description_label.setStyleSheet("font-weight: bold; color: #2c5aa0;")
        layout.addWidget(self.description_label)
        
        # Create table
        self.table = QTableWidget()
        self.setup_table()
        layout.addWidget(self.table)
        
        # Connect plan type changes
        self.plan_type_group.buttonClicked.connect(self.on_plan_type_changed)
        
        self.setLayout(layout)
    
    def setup_table(self):
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels([
            "Order", "Customer Item Name", "Customer Item Code", "Quantity", "Note"
        ])
        
        # Set column widths
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
    
    def on_plan_type_changed(self):
        descriptions = {
            self.plan_type_1: "Type 1: List of undelivered items by Customer, Order, Delivery Date",
            self.plan_type_2: "Type 2: List of undelivered items with customer by Customer, Order, Delivery date, Order date, Product",
            self.plan_type_3: "Type 3: Production Summary by Product, Month, Surface Treatment"
        }
        
        for button, description in descriptions.items():
            if button.isChecked():
                self.description_label.setText(description)
                break
    
    def generate_plan(self):
        """Generate the selected production plan"""
        if self.user and not self.permissions_manager.has_permission(self.user, "production_plans", "view"):
            QMessageBox.warning(self, "Permission Denied", "You don't have permission to generate production plans.")
            return
        
        # Determine plan type
        if self.plan_type_1.isChecked():
            plan_type = 'type1'
        elif self.plan_type_2.isChecked():
            plan_type = 'type2'
        else:
            plan_type = 'type3'
        
        # Show parameter dialog
        dialog = ProductionPlanDialog(plan_type, self.session, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            params = dialog.get_params()
            self.generate_plan_data(plan_type, params)
    
    def generate_plan_data(self, plan_type: str, params: dict):
        """Generate plan data based on type and parameters"""
        try:
            if plan_type == 'type1':
                data = self.generate_type1_plan(params)
            elif plan_type == 'type2':
                data = self.generate_type2_plan(params)
            elif plan_type == 'type3':
                data = self.generate_type3_plan(params)
            else:
                raise ValueError(f"Unknown plan type: {plan_type}")
            
            self.current_plan_data = data
            self.populate_table(plan_type, data)
            
            QMessageBox.information(self, "Plan Generated", f"Production plan {plan_type} has been generated successfully!")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error generating plan: {str(e)}")
    
    def generate_type1_plan(self, params: dict):
        """Generate Type 1 plan - undelivered items by customer, order, delivery date"""
        from models.database import Order, OrderItem, Customer, Item
        
        # Debug: print all parameters
        print(f"DEBUG: All parameters: {params}")
        
        # Build query
        query = self.session.query(OrderItem).join(Order).join(Customer).join(Item)
        
        # Debug: print initial query
        print(f"DEBUG: Initial query: {query}")
        
        # Apply filters
        if params.get('customer_id'):
            query = query.filter(Order.customer_id == params['customer_id'])
            print(f"DEBUG: After customer filter: {query}")
        
        if params.get('order_id'):
            query = query.filter(Order.id == params['order_id'])
            print(f"DEBUG: After order filter: {query}")
        
        # Filter by delivery date (specific date)
        if params.get('delivery_date'):
            query = query.filter(OrderItem.delivery_date == params['delivery_date'])
        
        # Filter by delivery date month/year
        delivery_months = params.get('delivery_months', [])
        delivery_year = params.get('delivery_year')
        if delivery_months and delivery_year:
            start_month = min(delivery_months)
            end_month = max(delivery_months)
            start_date = date(delivery_year, start_month, 1)
            if end_month == 12:
                end_date = date(delivery_year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(delivery_year, end_month + 1, 1) - timedelta(days=1)
            query = query.filter(OrderItem.delivery_date >= start_date, OrderItem.delivery_date <= end_date)
        
        # Filter for undelivered items
        query = query.filter(OrderItem.quantity > OrderItem.delivered_quantity)
        
        # Debug: let's see what the query looks like after undelivered filter
        print(f"DEBUG: Query after undelivered filter: {query}")
        
        order_items = query.all()
        
        # Debug: print some sample created_at dates
        print(f"DEBUG: Found {len(order_items)} items")
        for i, item in enumerate(order_items[:3]):  # Show first 3 items
            print(f"DEBUG: Item {i}: created_at = {item.created_at} (type: {type(item.created_at)})")
        
        # Format data
        data = []
        for item in order_items:
            # Format delivery date for display
            delivery_date_str = ''
            if item.delivery_date:
                delivery_date_str = item.delivery_date.strftime("%Y-%m-%d")
            
            data.append({
                'order': item.order.order_number,
                'customer_item_name': item.item.customer_item_name or '',
                'customer_item_code': item.item.customer_code,
                'delivery_date': delivery_date_str,
                'quantity': item.quantity - item.delivered_quantity,
                'note': ''
            })
        
        # Format delivery period for header/filename
        if delivery_months and delivery_year:
            if len(delivery_months) == 1:
                delivery_period = f"{delivery_months[0]:02d}-{delivery_year}"
            else:
                delivery_period = f"{min(delivery_months):02d}-{max(delivery_months):02d}-{delivery_year}"
        else:
            delivery_period = "all"
        
        # Get planned delivery date for header (most common delivery date)
        planned_delivery_date = None
        if order_items:
            delivery_dates = [item.delivery_date for item in order_items if item.delivery_date]
            if delivery_dates:
                # Get the most common delivery date
                from collections import Counter
                date_counts = Counter(delivery_dates)
                most_common_date = date_counts.most_common(1)[0][0]
                planned_delivery_date = most_common_date.strftime("%B %Y")  # e.g., "August 2025"
        
        return {
            'type': 'type1',
            'params': params,
            'data': data,
            'customer_name': order_items[0].order.customer.name_index if order_items else '',
            'delivery_date': delivery_period,
            'planned_delivery_date': planned_delivery_date,
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
    
    def generate_type2_plan(self, params: dict):
        """Generate Type 2 plan - undelivered items by customer, order, delivery date, order date, product"""
        from models.database import Order, OrderItem, Customer, Item, Product
        
        # Build query
        query = self.session.query(OrderItem).join(Order).join(Customer).join(Item).join(Product)
        
        # Apply filters
        if params.get('customer_id'):
            query = query.filter(Order.customer_id == params['customer_id'])
        
        if params.get('order_id'):
            query = query.filter(Order.id == params['order_id'])
        
        if params.get('order_date'):
            query = query.filter(Order.order_date == params['order_date'])
        
        if params.get('product_id'):
            query = query.filter(Item.product_id == params['product_id'])
        
        # Filter by delivery date month/year
        delivery_months = params.get('delivery_months', [])
        delivery_year = params.get('delivery_year')
        if delivery_months and delivery_year:
            start_month = min(delivery_months)
            end_month = max(delivery_months)
            start_date = date(delivery_year, start_month, 1)
            if end_month == 12:
                end_date = date(delivery_year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(delivery_year, end_month + 1, 1) - timedelta(days=1)
            query = query.filter(OrderItem.delivery_date >= start_date, OrderItem.delivery_date <= end_date)
        
        # Filter for undelivered items
        query = query.filter(OrderItem.quantity > OrderItem.delivered_quantity)
        
        order_items = query.all()
        
        # Format data
        data = []
        for item in order_items:
            data.append({
                'customer_name': item.order.customer.name_index,
                'order': item.order.order_number,
                'customer_item_name': item.item.customer_item_name or '',
                'customer_item_code': item.item.customer_code,
                'delivery_date': item.delivery_date.strftime("%Y-%m-%d"),
                'quantity': item.quantity - item.delivered_quantity,
                'note': ''
            })
        
        # Format delivery period for header/filename
        if delivery_months and delivery_year:
            if len(delivery_months) == 1:
                delivery_period = f"{delivery_months[0]:02d}-{delivery_year}"
            else:
                delivery_period = f"{min(delivery_months):02d}-{max(delivery_months):02d}-{delivery_year}"
        else:
            delivery_period = "all"
        
        return {
            'type': 'type2',
            'params': params,
            'data': data,
            'customer_name': order_items[0].order.customer.name_index if order_items else '',
            'delivery_date': delivery_period,
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
    
    def generate_type3_plan(self, params: dict):
        """Generate Type 3 plan - production summary by product, month, surface treatment"""
        from models.database import Order, OrderItem, Customer, Item, Product
        
        # Build query
        query = self.session.query(OrderItem).join(Order).join(Customer).join(Item).join(Product)
        
        # Apply filters
        if params.get('customer_id'):
            query = query.filter(Order.customer_id == params['customer_id'])
        
        if params.get('surface_treatment'):
            query = query.filter(OrderItem.surface_treatment == params['surface_treatment'])
        
        # Filter by delivery date month/year
        delivery_months = params.get('delivery_months', [])
        delivery_year = params.get('delivery_year')
        if delivery_months and delivery_year:
            start_month = min(delivery_months)
            end_month = max(delivery_months)
            start_date = date(delivery_year, start_month, 1)
            if end_month == 12:
                end_date = date(delivery_year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(delivery_year, end_month + 1, 1) - timedelta(days=1)
            query = query.filter(OrderItem.delivery_date >= start_date, OrderItem.delivery_date <= end_date)
        
        # Filter for undelivered items
        query = query.filter(OrderItem.quantity > OrderItem.delivered_quantity)
        
        order_items = query.all()
        
        # Group by product, month, and surface treatment
        product_data = {}
        month_names = {
            1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun',
            7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'
        }
        
        # Track monthly totals for subtotals
        monthly_totals = {}
        
        for item in order_items:
            product_name = item.item.product.name
            surface_treatment = item.surface_treatment or 'KATAFOREZA'
            quantity = item.quantity - item.delivered_quantity
            
            # Get delivery month
            delivery_month = item.delivery_date.month if item.delivery_date else None
            month_name = month_names.get(delivery_month, 'Unknown') if delivery_month else 'Unknown'
            
            if product_name not in product_data:
                product_data[product_name] = {}
            
            # Create column key: "KATAFOREZA Aug", "FOSFAT Sep", etc.
            column_key = f"{surface_treatment} {month_name}"
            
            if column_key not in product_data[product_name]:
                product_data[product_name][column_key] = 0
            
            product_data[product_name][column_key] += quantity
            
            # Track monthly totals
            if month_name not in monthly_totals:
                monthly_totals[month_name] = 0
            monthly_totals[month_name] += quantity
        
        # Ensure we have monthly totals even if no data was found
        if not monthly_totals and product_data:
            # If no monthly totals but we have product data, calculate from product data
            for product_treatments in product_data.values():
                for column_key, quantity in product_treatments.items():
                    # Extract month from column key (e.g., "KATAFOREZA Aug" -> "Aug")
                    parts = column_key.split()
                    if len(parts) >= 2:
                        month_name = parts[-1]  # Last part is the month
                        if month_name not in monthly_totals:
                            monthly_totals[month_name] = 0
                        monthly_totals[month_name] += quantity
        
        # Add TOTAL columns for each month to each product
        for product_name in product_data:
            for month_name in monthly_totals:
                total_column_key = f"TOTAL {month_name}"
                # Calculate total for this product in this month
                product_month_total = 0
                for column_key, quantity in product_data[product_name].items():
                    if month_name in column_key:
                        product_month_total += quantity
                product_data[product_name][total_column_key] = product_month_total
        
        # Format delivery period for header/filename
        if delivery_months and delivery_year:
            if len(delivery_months) == 1:
                delivery_period = f"{delivery_months[0]:02d}-{delivery_year}"
            else:
                delivery_period = f"{min(delivery_months):02d}-{max(delivery_months):02d}-{delivery_year}"
        else:
            delivery_period = "all"
        
        return {
            'type': 'type3',
            'params': params,
            'data': product_data,
            'monthly_totals': monthly_totals,
            'customer_name': order_items[0].order.customer.name_index if order_items else '',
            'delivery_date': delivery_period,
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
    
    def populate_table(self, plan_type: str, data: dict):
        """Populate the table with plan data"""
        if plan_type in ['type1', 'type2']:
            self.setup_type1_2_table()
            items = data.get('data', [])
            
            # Sort items for Type 2 to match Excel export
            if plan_type == 'type2':
                items = sorted(items, key=lambda x: x.get('customer_name', ''))
            
            self.table.setRowCount(len(items))
            
            for row, item in enumerate(items):
                if plan_type == 'type1':
                    self.table.setItem(row, 0, QTableWidgetItem(item['order']))
                    self.table.setItem(row, 1, QTableWidgetItem(item['customer_item_name']))
                    self.table.setItem(row, 2, QTableWidgetItem(item['customer_item_code']))
                    self.table.setItem(row, 3, QTableWidgetItem(item['delivery_date']))
                    self.table.setItem(row, 4, QTableWidgetItem(str(item['quantity'])))
                    self.table.setItem(row, 5, QTableWidgetItem(item['note']))
                else:  # type2
                    self.table.setItem(row, 0, QTableWidgetItem(item['customer_name']))
                    self.table.setItem(row, 1, QTableWidgetItem(item['order']))
                    self.table.setItem(row, 2, QTableWidgetItem(item['customer_item_name']))
                    self.table.setItem(row, 3, QTableWidgetItem(item['customer_item_code']))
                    self.table.setItem(row, 4, QTableWidgetItem(item['delivery_date']))
                    self.table.setItem(row, 5, QTableWidgetItem(str(item['quantity'])))
                    self.table.setItem(row, 6, QTableWidgetItem(item['note']))
        
        elif plan_type == 'type3':
            self.setup_type3_table()
            product_data = data.get('data', {})
            products = list(product_data.keys())
            
            # Get all column keys and sort them by month, then by treatment type
            all_columns = set()
            for product_data_dict in product_data.values():
                all_columns.update(product_data_dict.keys())
            
            # Sort columns by month first, then by treatment type
            def sort_columns(column_key):
                # Extract month and treatment from column key
                parts = column_key.split()
                if len(parts) >= 2:
                    treatment = ' '.join(parts[:-1])  # Everything except last part
                    month = parts[-1]  # Last part is month
                    
                    # Define treatment order
                    treatment_order = {
                        'FOSFAT': 1,
                        'KATAFOREZA': 2, 
                        'ZINEK': 3,
                        'TOTAL': 4
                    }
                    
                    # Get treatment priority (default to 999 for unknown treatments)
                    treatment_priority = treatment_order.get(treatment, 999)
                    
                    # Return tuple for sorting: (month, treatment_priority, column_key)
                    return (month, treatment_priority, column_key)
                else:
                    return (column_key, 999, column_key)
            
            all_columns = sorted(list(all_columns), key=sort_columns)
            
            # Set up table
            self.table.setRowCount(len(products))
            
            # Set up columns
            self.table.setColumnCount(1 + len(all_columns) + 1)  # Product + all columns + total
            headers = ["Product"] + all_columns + ["Total"]
            self.table.setHorizontalHeaderLabels(headers)
            
            # Set header background colors
            from PyQt6.QtGui import QColor, QBrush
            for col, header_text in enumerate(headers):
                header_item = self.table.horizontalHeaderItem(col)
                if header_text.startswith('TOTAL'):
                    header_item.setBackground(QBrush(QColor(173, 216, 230)))  # Light blue
                elif header_text == "Total":
                    header_item.setBackground(QBrush(QColor(100, 149, 237)))  # Darker blue
            
            # Populate product rows
            for row, product in enumerate(products):
                self.table.setItem(row, 0, QTableWidgetItem(product))
                
                total = 0
                for col, column_key in enumerate(all_columns, 1):
                    quantity = product_data[product].get(column_key, 0)
                    item = QTableWidgetItem(str(quantity))
                    
                    # Set background color for TOTAL columns (light blue)
                    if column_key.startswith('TOTAL'):
                        from PyQt6.QtGui import QColor, QBrush
                        item.setBackground(QBrush(QColor(173, 216, 230)))  # Light blue
                    
                    self.table.setItem(row, col, item)
                    total += quantity
                
                # Set background color for the final Total column (darker blue)
                total_item = QTableWidgetItem(str(total))
                from PyQt6.QtGui import QColor, QBrush
                total_item.setBackground(QBrush(QColor(100, 149, 237)))  # Darker blue
                self.table.setItem(row, len(all_columns) + 1, total_item)
        
        # For Type 2, highlight first occurrence of each customer in the table
        if plan_type == 'type2':
            from PyQt6.QtGui import QColor, QBrush
            print(f"DEBUG: Starting highlighting for Type 2 plan with {self.table.rowCount()} rows")
            seen_customers = set()
            rows_to_highlight = []
            
            # First pass: identify which rows to highlight
            for row in range(self.table.rowCount()):
                customer_item = self.table.item(row, 0)  # Customer is in first column
                if customer_item:
                    customer_name = customer_item.text()
                    print(f"DEBUG: Row {row}, Customer: {customer_name}")
                    if customer_name and customer_name not in seen_customers:
                        print(f"DEBUG: First occurrence of {customer_name} at row {row}")
                        # Highlight the row AFTER the first occurrence (row + 1)
                        if row < self.table.rowCount() - 1:
                            rows_to_highlight.append(row + 1)
                            print(f"DEBUG: Highlighting row {row + 1} (after first occurrence)")
                        seen_customers.add(customer_name)
            
            print(f"DEBUG: Rows to highlight: {rows_to_highlight}")
            
            # Second pass: apply highlighting with more visible color
            for row in rows_to_highlight:
                print(f"DEBUG: Processing row {row}")
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    if item:
                        print(f"DEBUG: Found item at row {row}, col {col}: {item.text()}")
                        # Use QBrush instead of QColor directly
                        yellow_brush = QBrush(QColor(255, 255, 0))
                        item.setBackground(yellow_brush)
                        print(f"DEBUG: Set yellow background for row {row}, col {col}")
                    else:
                        print(f"DEBUG: No item found at row {row}, col {col}")
            
            # Force table to update and repaint
            self.table.viewport().update()
            self.table.repaint()
            print(f"DEBUG: Highlighting complete")
        
        # Resize columns
        self.table.resizeColumnsToContents()
    
    def setup_type1_2_table(self):
        """Setup table for type 1 and 2 plans"""
        if self.plan_type_1.isChecked():
            self.table.setColumnCount(6)
            self.table.setHorizontalHeaderLabels([
                "Order", "Customer Item Name", "Customer Item Code", "Delivery Date", "Quantity", "Note"
            ])
        else:  # type2
            self.table.setColumnCount(7)
            self.table.setHorizontalHeaderLabels([
                "Customer", "Order", "Customer Item Name", "Customer Item Code", "Delivery Date", "Quantity", "Note"
            ])
    
    def setup_type3_table(self):
        """Setup table for type 3 plan"""
        # This will be set up dynamically based on available surface treatments
        pass
    
    def export_plan(self):
        """Export the current plan to Excel with header information"""
        if not self.current_plan_data:
            QMessageBox.warning(self, "No Plan", "Please generate a plan first.")
            return
        
        try:
            data = self.current_plan_data
            plan_type = data['type']
            
            # Create header information based on plan type
            header_info = []
            
            if plan_type == 'type1':
                # Type 1: Customer name index, planned delivery date, current timestamp
                header_info.append(f"Customer: {data.get('customer_name', 'N/A')}")
                if data.get('planned_delivery_date'):
                    header_info.append(f"Planned Delivery Date: {data.get('planned_delivery_date', 'N/A')}")
                else:
                    header_info.append(f"Delivery Period: {data.get('delivery_date', 'N/A')}")
                header_info.append(f"Generated: {data.get('timestamp', 'N/A')}")
            elif plan_type in ['type2', 'type3']:
                # Type 2 & 3: Current timestamp, month/year of delivery date
                header_info.append(f"Generated: {data.get('timestamp', 'N/A')}")
                header_info.append(f"Delivery Period: {data.get('delivery_date', 'N/A')}")
                if plan_type == 'type2':
                    header_info.append(f"Customer: {data.get('customer_name', 'N/A')}")
            
            # Create the data DataFrame
            if plan_type in ['type1', 'type2']:
                df = pd.DataFrame(data.get('data', []))
                # Sort by customer for type2, by order for type1
                if plan_type == 'type2' and not df.empty and 'customer_name' in df.columns:
                    df = df.sort_values('customer_name').reset_index(drop=True)
                elif plan_type == 'type1' and not df.empty and 'order' in df.columns:
                    df = df.sort_values('order').reset_index(drop=True)
            elif plan_type == 'type3':
                product_data = data.get('data', {})
                # Convert to DataFrame format with dynamic columns including TOTAL columns
                rows = []
                all_columns = set()
                for product, treatments in product_data.items():
                    all_columns.update(treatments.keys())
                
                # Sort columns by month first, then by treatment type
                def sort_columns(column_key):
                    # Extract month and treatment from column key
                    parts = column_key.split()
                    if len(parts) >= 2:
                        treatment = ' '.join(parts[:-1])  # Everything except last part
                        month = parts[-1]  # Last part is month
                        
                        # Define treatment order
                        treatment_order = {
                            'FOSFAT': 1,
                            'KATAFOREZA': 2, 
                            'ZINEK': 3,
                            'TOTAL': 4
                        }
                        
                        # Get treatment priority (default to 999 for unknown treatments)
                        treatment_priority = treatment_order.get(treatment, 999)
                        
                        # Return tuple for sorting: (month, treatment_priority, column_key)
                        return (month, treatment_priority, column_key)
                    else:
                        return (column_key, 999, column_key)
                
                all_columns = sorted(list(all_columns), key=sort_columns)
                
                # Add product rows
                for product, treatments in product_data.items():
                    row = {'Product': product}
                    total = 0
                    for column_key in all_columns:
                        quantity = treatments.get(column_key, 0)
                        row[column_key] = quantity
                        total += quantity
                    row['Total'] = total
                    rows.append(row)
                
                df = pd.DataFrame(rows)
                # Sort by product name
                if not df.empty:
                    df = df.sort_values('Product')
            
            # Save to file with header information
            # For type 2 and 3, check if a specific customer was selected in the dialog
            if plan_type in ['type2', 'type3']:
                customer_id = data.get('params', {}).get('customer_id')
                if customer_id is None:
                    customer_part = "all_customers"
                else:
                    customer_part = data.get('customer_name', 'all').replace(' ', '_')
            else:
                customer_part = data.get('customer_name', 'all').replace(' ', '_')
            
            delivery_part = data.get('delivery_date', 'all').replace('/', '-')
            filename, _ = QFileDialog.getSaveFileName(
                self, "Export Production Plan", 
                f"Production_Plan_{plan_type}_{customer_part}_{delivery_part}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if filename:
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    # Create header row with all information
                    header_row = []
                    if plan_type == 'type1':
                        if data.get('planned_delivery_date'):
                            header_row = [f"Customer: {data.get('customer_name', 'N/A')}", 
                                        f"Planned Delivery Date: {data.get('planned_delivery_date', 'N/A')}", 
                                        f"Generated: {data.get('timestamp', 'N/A')}"]
                        else:
                            header_row = [f"Customer: {data.get('customer_name', 'N/A')}", 
                                        f"Delivery Period: {data.get('delivery_date', 'N/A')}", 
                                        f"Generated: {data.get('timestamp', 'N/A')}"]
                    elif plan_type == 'type2':
                        header_row = [f"Generated: {data.get('timestamp', 'N/A')}", 
                                    f"Delivery Period: {data.get('delivery_date', 'N/A')}", 
                                    f"Customer: {data.get('customer_name', 'N/A')}"]
                    elif plan_type == 'type3':
                        header_row = [f"Generated: {data.get('timestamp', 'N/A')}", 
                                    f"Delivery Period: {data.get('delivery_date', 'N/A')}"]
                    
                    # Write header row
                    header_df = pd.DataFrame([header_row])
                    header_df.to_excel(writer, sheet_name='Production Plan', index=False, header=False)
                    
                    # Write the main data directly after header
                    if not df.empty:
                        df.to_excel(writer, sheet_name='Production Plan', startrow=1, index=False)
                    else:
                        # If no data, just write a message
                        no_data_df = pd.DataFrame({'No Data': ['No items found for the selected criteria']})
                        no_data_df.to_excel(writer, sheet_name='Production Plan', startrow=1, index=False)
                    
                    # Auto-adjust column widths and add borders
                    worksheet = writer.sheets['Production Plan']
                    from openpyxl.styles import Border, Side
                    
                    # Define border style
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Auto-adjust column widths and add borders to data cells
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)  # Add some padding, max 50
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Add borders to all cells except the header row (row 1)
                    for row in worksheet.iter_rows(min_row=2):
                        for cell in row:
                            cell.border = thin_border
                    
                    # For Type 2, highlight first occurrence of each customer
                    if plan_type == 'type2' and not df.empty:
                        from openpyxl.styles import PatternFill
                        customer_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                        
                        # Find first occurrence of each customer and highlight the row AFTER it
                        seen_customers = set()
                        for row_idx, row_data in df.iterrows():
                            customer_name = row_data.get('customer_name', '')
                            if customer_name and customer_name not in seen_customers:
                                # Highlight the row AFTER the first occurrence (row_idx + 1)
                                excel_row = row_idx + 3  # +3 because Excel is 1-based, data starts at row 2 (after header), so +2 for data row, +1 for next row
                                if excel_row <= len(df) + 2:  # Make sure we don't exceed the data range
                                    for col in range(1, len(df.columns) + 1):
                                        cell = worksheet.cell(row=excel_row, column=col)
                                        cell.fill = customer_fill
                                seen_customers.add(customer_name)
                    
                    # For Type 3, add background colors for TOTAL columns
                    if plan_type == 'type3' and not df.empty:
                        from openpyxl.styles import PatternFill
                        
                        # Define colors
                        light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                        darker_blue_fill = PatternFill(start_color="6495ED", end_color="6495ED", fill_type="solid")
                        
                        # Get column headers to identify TOTAL columns
                        headers = list(df.columns)
                        
                        # Apply colors to header row (row 2 after header info)
                        for col_idx, header in enumerate(headers, 1):
                            cell = worksheet.cell(row=2, column=col_idx)
                            if header.startswith('TOTAL'):
                                cell.fill = light_blue_fill
                            elif header == "Total":
                                cell.fill = darker_blue_fill
                        
                        # Apply colors to data rows
                        for row_idx in range(3, len(df) + 3):  # Start from row 3 (after headers)
                            for col_idx, header in enumerate(headers, 1):
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                if header.startswith('TOTAL'):
                                    cell.fill = light_blue_fill
                                elif header == "Total":
                                    cell.fill = darker_blue_fill
                
                QMessageBox.information(self, "Export Successful", f"Plan exported to {filename}")
        
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Error exporting plan: {str(e)}") 